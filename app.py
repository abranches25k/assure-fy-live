# --- IMPORTS NECESSÁRIOS ---
from flask import Flask, render_template, request, redirect, url_for, flash
from flask_sqlalchemy import SQLAlchemy
from apscheduler.schedulers.background import BackgroundScheduler
from datetime import datetime, timedelta, date
import os
import atexit
import requests 
from werkzeug.utils import secure_filename
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import load_workbook 
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
import smtplib
from email.message import EmailMessage

# --- CONFIGURAÇÃO INICIAL DO FLASK ---
app = Flask(__name__)

# V40.0: Configuração final do caminho do banco de dados (usa /tmp no Render)
db_path = os.environ.get("DATABASE_URL", "sqlite:///saas.db")
# Se estivermos no Render, forçamos o DB para /tmp/saas.db por questões de permissão
if os.environ.get('RENDER', '') == 'true':
    db_path = "sqlite:////tmp/saas.db"
    
app.config['SQLALCHEMY_DATABASE_URI'] = db_path
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['LOGO_FOLDER'] = 'logos'

# V34.0: Usar variável de ambiente (SECRET_KEY)
app.config['SECRET_KEY'] = os.environ.get("SECRET_KEY", "SUA_CHAVE_SECRETA_MUITO_LONGA_E_COMPLEXA") 
app.secret_key = app.config['SECRET_KEY'] 

db = SQLAlchemy(app)

# --- CONFIGURAÇÃO DO LOGIN ---
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login' 

# --- CONFIGURAÇÃO DE E-MAIL E TELEGRAM (COM VARIÁVEIS DE AMBIENTE) ---
EMAIL_USER = os.environ.get("EMAIL_USER", "qualify.relatorios@gmail.com") 
EMAIL_PASS = os.environ.get("EMAIL_PASS", "ukwmtfberzqafswj") 
SMTP_SERVER = os.environ.get("SMTP_SERVER", "smtp.gmail.com")
SMTP_PORT = int(os.environ.get("SMTP_PORT", 587))
ADMIN_EMAIL = os.environ.get("ADMIN_EMAIL", "qualify.relatorios@gmail.com")
TELEGRAM_BOT_TOKEN = os.environ.get("TELEGRAM_BOT_TOKEN", "SEU_TOKEN_DO_BOTFATHER_AQUI")

# --- MODELO DE USUÁRIO ---
class Usuario(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(100), unique=True, nullable=False)
    senha_hash = db.Column(db.String(200), nullable=False)
    nome_empresa = db.Column(db.String(100), nullable=False)
    logo_path = db.Column(db.String(255), nullable=True) 
    limite_tarefas = db.Column(db.Integer, default=3) 
    executions_count = db.Column(db.Integer, default=0) 

    tarefas = db.relationship('Tarefa', backref='owner', lazy='dynamic')
    logs = db.relationship('LogExecucao', backref='log_owner', lazy='dynamic')
    historico = db.relationship('HistoricoProducao', backref='hist_user_owner', lazy='dynamic') 

    def set_password(self, senha):
        self.senha_hash = generate_password_hash(senha)

    def check_password(self, senha):
        return check_password_hash(self.senha_hash, senha)

@login_manager.user_loader
def load_user(user_id):
    return Usuario.query.get(int(user_id))

# --- NOVO MODELO: HISTÓRICO DE PRODUÇÃO ---
class HistoricoProducao(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    data_registro = db.Column(db.Date, default=date.today)
    producao_total = db.Column(db.Float, nullable=False)
    ticket_medio = db.Column(db.Float, nullable=False)
    
    tarefa_id = db.Column(db.Integer, db.ForeignKey('tarefa.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    
    __table_args__ = (db.UniqueConstraint('data_registro', 'tarefa_id', name='_data_tarefa_uc'),)

# --- MODELO DA TAREFA ---
class Tarefa(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome_cliente = db.Column(db.String(100), nullable=False)
    caminho_arquivo = db.Column(db.String(200), nullable=False)
    hora_agendamento = db.Column(db.String(5), nullable=False) 
    email_destino = db.Column(db.String(100), nullable=False) 
    
    filtro_tipo = db.Column(db.String(10), nullable=False, default='DATA') 
    coluna_data = db.Column(db.Integer, nullable=True)
    dias_alerta = db.Column(db.Integer, nullable=True) 
    coluna_status = db.Column(db.Integer, nullable=True) 
    palavra_chave = db.Column(db.String(100), nullable=True)
    telegram_chat_id = db.Column(db.String(50), nullable=True) 
    
    coluna_valor = db.Column(db.Integer, nullable=True) 
    coluna_id_unico = db.Column(db.Integer, nullable=True) 
    
    coluna_base = db.Column(db.Integer, nullable=True)
    coluna_acionado = db.Column(db.Integer, nullable=True)
    coluna_atendido = db.Column(db.Integer, nullable=True)
    coluna_cpc = db.Column(db.Integer, nullable=True)
    coluna_acordos = db.Column(db.Integer, nullable=True)
    coluna_pagamentos = db.Column(db.Integer, nullable=True)

    user_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False) 
    logs = db.relationship('LogExecucao', backref='log_tarefa', lazy='dynamic')
    
    historico = db.relationship('HistoricoProducao', backref='producao_tarefa', lazy='dynamic')

    def __repr__(self):
        return f'<Tarefa {self.nome_cliente}>'

# --- MODELO: LOG DE EXECUÇÃO ---
class LogExecucao(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)
    status = db.Column(db.String(10), nullable=False)
    mensagem = db.Column(db.Text, nullable=True)
    
    tarefa_id = db.Column(db.Integer, db.ForeignKey('tarefa.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    
    def __repr__(self):
        return f'<Log {self.status} {self.timestamp}>'


# --- FUNÇÕES CORE ---
def enviar_email_com_anexo(destino, assunto, corpo, anexo_path):
    smtp_error_message = None 
    try:
        msg = EmailMessage()
        msg['Subject'] = assunto
        msg['From'] = EMAIL_USER
        msg['To'] = destino
        msg.set_content(corpo)
        
        if os.path.exists(anexo_path):
            with open(anexo_path, 'rb') as f:
                file_data = f.read()
                file_name = os.path.basename(anexo_path)
            msg.add_attachment(file_data, maintype='application', subtype='octet-stream', filename=file_name)

        # V31.0: Usar conexão SMTP padrão (STARTTLS)
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT, timeout=5) as server:
            server.ehlo()
            server.starttls() 
            server.login(EMAIL_USER, EMAIL_PASS.replace(' ', '')) 
            server.send_message(msg)
            
        return True
    except Exception as e:
        smtp_error_message = f"SMTP ERROR: {type(e).__name__} - {str(e)[:150]}"
        print(f"\n\n🚨 ERRO SMTP CAPTURADO: {smtp_error_message}\n\n")
        
        return False

def enviar_telegram(chat_id, mensagem):
    """Envia uma mensagem de alerta via API do Telegram."""
    if not TELEGRAM_BOT_TOKEN or not chat_id:
        return False
        
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
    payload = {
        "chat_id": chat_id,
        "text": mensagem,
        "parse_mode": "Markdown"
    }
    try:
        response = requests.post(url, data=payload)
        return response.json().get("ok", False)
    except Exception as e:
        print(f"Erro ao enviar Telegram: {e}")
        return False

def processar_e_enviar_relatorio(tarefa_id):
    with app.app_context():
        tarefa = Tarefa.query.get(tarefa_id)
        if not tarefa: return
        user = Usuario.query.get(tarefa.user_id)
        if not user: return

        if user.limite_tarefas == 3:
            if user.executions_count >= 5:
                novo_log = LogExecucao(status="BLOQ", mensagem=f"Bloqueado: Limite de 5 execuções de testes gratuitos atingido.", tarefa_id=tarefa.id, user_id=user.id)
                db.session.add(novo_log); db.session.commit()
                return 
            
            user.executions_count += 1
            db.session.commit()
        
        pdf_caminho = None
        status = "SUCESSO"
        mensagem = ""
        workbook = None 
        smtp_falhou = False 

        try:
            # 1. Leitura e Processamento (Lógica Mantida)
            nome_pdf = f"Relatorio_{tarefa_id}_{tarefa.nome_cliente}.pdf"
            pdf_caminho = os.path.join(app.config['UPLOAD_FOLDER'], nome_pdf)
            dados_relatorio = ""
            
            workbook = load_workbook(tarefa.caminho_arquivo)
            planilha = workbook.active 

            itens_filtrados = []
            
            # --- LÓGICA DE FILTRO CONDICIONAL (V8.0) --- (Mantida)
            if tarefa.filtro_tipo == 'DATA':
                hoje = datetime.now().date()
                limite_alerta = hoje + timedelta(days=tarefa.dias_alerta)
                coluna_monitorada = tarefa.coluna_data
                for i in range(2, planilha.max_row + 1):
                    id_processo = planilha.cell(row=i, column=1).value 
                    data_celula = planilha.cell(row=i, column=coluna_monitorada).value 
                    data_prazo = None
                    if isinstance(data_celula, datetime): data_prazo = data_celula.date()
                    elif isinstance(data_celula, str):
                        try: data_prazo = datetime.strptime(data_celula, '%d/%m/%Y').date()
                        except ValueError: continue 
                    if data_prazo and hoje <= data_prazo <= limite_alerta:
                        itens_filtrados.append({"id": id_processo, "data_ou_status": data_prazo.strftime('%d/%m/%Y')})
                dados_relatorio += f"RELATÓRIO DE PRAZOS CRÍTICOS (Regra: DATA)\n\n"
                dados_relatorio += f"- Coluna Monitorada: {coluna_monitorada}\n"
                dados_relatorio += f"- Alerta Ativo para: {tarefa.dias_alerta} dias\n"

            elif tarefa.filtro_tipo == 'STATUS':
                palavra_alvo = tarefa.palavra_chave.strip().upper()
                coluna_monitorada = tarefa.coluna_status
                for i in range(2, planilha.max_row + 1):
                    id_processo = planilha.cell(row=i, column=1).value 
                    status_celula = planilha.cell(row=i, column=coluna_monitorada).value 
                    if status_celula and str(status_celula).strip().upper() == palavra_alvo:
                        itens_filtrados.append({"id": id_processo, "data_ou_status": status_celula})
                dados_relatorio += f"RELATÓRIO DE STATUS (Regra: STATUS)\n\n"
                dados_relatorio += f"- Coluna Monitorada: {coluna_monitorada}\n"
                dados_relatorio += f"- Buscando pela Palavra: {tarefa.palavra_chave}\n"
            
            elif tarefa.filtro_tipo == 'FUNIL':
                base_total = planilha.max_row - 1
                acionado = 0
                atendido = 0
                cpc = 0
                acordos = 0
                pagamentos = 0
                for i in range(2, planilha.max_row + 1):
                    if tarefa.coluna_acionado and planilha.cell(row=i, column=tarefa.coluna_acionado).value: acionado += 1
                    if tarefa.coluna_atendido and planilha.cell(row=i, column=tarefa.coluna_atendido).value: atendido += 1
                    if tarefa.coluna_cpc and planilha.cell(row=i, column=tarefa.coluna_cpc).value: cpc += 1
                    if tarefa.coluna_acordos and planilha.cell(row=i, column=tarefa.coluna_acordos).value: acordos += 1
                    if tarefa.coluna_pagamentos and planilha.cell(row=i, column=tarefa.coluna_pagamentos).value: pagamentos += 1
                perc_acordos = (acordos / cpc) * 100 if cpc > 0 else 0
                perc_pagamentos = (pagamentos / acordos) * 100 if acordos > 0 else 0
                dados_relatorio += f"--- ANÁLISE DE FUNIL (Batimento de Carga) ---\n"
                dados_relatorio += f"Base Total de Clientes: {base_total}\n"
                dados_relatorio += f"1. Acionado: {acionado}\n"
                dados_relatorio += f"2. Atendido: {atendido}\n"
                dados_relatorio += f"3. CPC (Contato Positivo): {cpc}\n"
                dados_relatorio += f"4. Acordos Formalizados: {acordos}\n"
                dados_relatorio += f"5. Pagamentos Recebidos: {pagamentos}\n"
                dados_relatorio += f"\n--- TAXAS DE CONVERSÃO ---\n"
                dados_relatorio += f"Acordos/CPC: {perc_acordos:.2f}%\n"
                dados_relatorio += f"Pagamentos/Acordos: {perc_pagamentos:.2f}%\n"
                dados_relatorio += "---------------------------------------\n"


            # 2. CÁLCULO BI HISTÓRICO (V22.0)
            if tarefa.coluna_valor and tarefa.coluna_id_unico:
                
                producao_total_hoje = 0.0
                itens_unicos_hoje = set()
                
                for i in range(2, planilha.max_row + 1):
                    valor_celula = planilha.cell(row=i, column=tarefa.coluna_valor).value
                    id_unico_celula = planilha.cell(row=i, column=tarefa.coluna_id_unico).value

                    try: valor = float(valor_celula) if valor_celula else 0.0
                    except (ValueError, TypeError): valor = 0.0

                    if valor > 0:
                        producao_total_hoje += valor
                        itens_unicos_hoje.add(id_unico_celula)

                num_itens_unicos = len(itens_unicos_hoje)
                ticket_medio_hoje = producao_total_hoje / num_itens_unicos if num_itens_unicos > 0 else 0.0

                
                ontem = date.today() - timedelta(days=1)
                historico_ontem = HistoricoProducao.query.filter_by(tarefa_id=tarefa.id, data_registro=ontem).first()
                
                if historico_ontem:
                    diff_producao = producao_total_hoje - historico_ontem.producao_total
                    diff_ticket = ticket_medio_hoje - historico_ontem.ticket_medio
                    
                    perc_producao = (diff_producao / historico_ontem.producao_total) * 100 if historico_ontem.producao_total != 0 else 0
                    perc_ticket = (diff_ticket / historico_ontem.ticket_medio) * 100 if historico_ontem.ticket_medio != 0 else 0

                    dados_relatorio += f"\n--- RELATÓRIO BI (DIA-A-DIA) ---\n"
                    dados_relatorio += f"Produção Total Hoje: R$ {producao_total_hoje:.2f}\n"
                    dados_relatorio += f"Ticket Médio Hoje: R$ {ticket_medio_hoje:.2f}\n"
                    dados_relatorio += f"\n--- COMPARAÇÃO COM ONTEM ---\n"
                    dados_relatorio += f"Diferença Produção: R$ {diff_producao:+.2f} ({perc_producao:+.2f}%)\n"
                    dados_relatorio += f"Diferença Ticket Médio: R$ {diff_ticket:+.2f} ({perc_ticket:+.2f}%)\n"
                    dados_relatorio += f"---------------------------------\n\n"
                
                novo_historico = HistoricoProducao(
                    data_registro=date.today(),
                    producao_total=producao_total_hoje,
                    ticket_medio=ticket_medio_hoje,
                    tarefa_id=tarefa.id,
                    user_id=user.id
                )
                
                HistoricoProducao.query.filter_by(tarefa_id=tarefa.id, data_registro=date.today()).delete()
                db.session.add(novo_historico)
                db.session.commit()
            
            if tarefa.filtro_tipo != 'FUNIL':
                dados_relatorio += f"- Total de Itens em Risco/Encontrados: {len(itens_filtrados)}\n"
                dados_relatorio += "="*50 + "\n"
                
                if itens_filtrados:
                    dados_relatorio += "\nITENS ENCONTRADOS:\n"
                    for item in itens_filtrados:
                        dados_relatorio += f"  > Item ID {item['id']} ({item['data_ou_status']})\n"
                else:
                    dados_relatorio += "\n** PARABÉNS! Nenhum item encontrado pela regra. **\n"

            email_success = enviar_email_com_anexo(tarefa.email_destino, f"Relatório ASSURE FY: {tarefa.nome_cliente}", dados_relatorio, pdf_caminho)
            telegram_success = enviar_telegram(tarefa.telegram_chat_id, f"🚨 ALERTA ASSURE FY: Relatório {tarefa.nome_cliente} executado.")
            
            if not email_success:
                smtp_falhou = True 
            
        except Exception as e:
            status = "FALHA"
            mensagem = f"ERRO CRÍTICO na Tarefa {tarefa.id}: {str(e)[:500]}"
            
        finally:
            if 'workbook' in locals() and workbook:
                try: del workbook 
                except: pass

            if smtp_falhou:
                mensagem_log = "ERRO SMTP: Falha ao conectar/autenticar no servidor de e-mail. Verifique a Senha de App."
                
                novo_log_email = LogExecucao(
                    status="SMTP",
                    mensagem=mensagem_log,
                    tarefa_id=tarefa.id,
                    user_id=user.id
                )
                db.session.add(novo_log_email)
                
                if status == "SUCESSO": status = "AVISO"
                
            novo_log = LogExecucao(
                status=status,
                mensagem=mensagem,
                tarefa_id=tarefa.id,
                user_id=user.id
            )
            db.session.add(novo_log)
            db.session.commit()

            if pdf_caminho and os.path.exists(pdf_caminho):
                os.remove(pdf_caminho)

def agendar_tarefa_core(tarefa_id, hora_str):
    try: scheduler.remove_job(str(tarefa_id))
    except: pass 

    hora, minuto = map(int, hora_str.split(':'))
    
    scheduler.add_job(
        func=processar_e_enviar_relatorio,
        trigger='cron',
        hour=hora, minute=minuto, id=str(tarefa_id), args=[tarefa_id]
    )
    print(f"Tarefa {tarefa_id} agendada para {hora_str} diariamente.")


# --- ROTAS DE AUTENTICAÇÃO E ADMIN (MANTIDAS) ---
@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        email = request.form.get('email')
        senha = request.form.get('senha')
        nome_empresa = request.form.get('nome_empresa')

        user = Usuario.query.filter_by(email=email).first()

        if user:
            flash("Email já registrado. Tente outro ou faça login.", 'danger')
            return redirect(url_for('register'))

        novo_usuario = Usuario(email=email, nome_empresa=nome_empresa, limite_tarefas=3, executions_count=0)
        novo_usuario.set_password(senha)
        
        db.session.add(novo_usuario)
        db.session.commit()
        
        login_user(novo_usuario)
        flash("Conta criada com sucesso! Você está no Plano TESTE GRATUITO (3 Tarefas / 5 Execuções).", 'message')
        return redirect(url_for('planos')) 

    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('lista_tarefas'))
        
    if request.method == 'POST':
        email = request.form.get('email')
        senha = request.form.get('senha')
        
        user = Usuario.query.filter_by(email=email).first()

        if user and user.check_password(senha):
            login_user(user)
            
            if user.limite_tarefas == 3:
                flash(f"Bem-vindo de volta! Você está no Teste Gratuito. Execuções restantes: {5 - user.executions_count}.", 'warning')
                return redirect(url_for('planos'))
                
            return redirect(url_for('lista_tarefas'))
        else:
            flash("Email ou Senha Inválidos.", 'danger')
            return redirect(url_for('login'))
    
    return render_template('login.html')

# --- CORREÇÃO DO LOGOUT (MANTIDA) ---
@app.route('/logout')
@login_required
def logout():
    logout_user() 
    flash("Você foi desconectado com sucesso.", 'message')
    return redirect(url_for('login')) 


ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg'}
def allowed_file(filename):
    return '.' in filename and \
            filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/perfil', methods=['GET', 'POST'])
@login_required
def perfil():
    if not os.path.exists(app.config['LOGO_FOLDER']):
        os.makedirs(app.config['LOGO_FOLDER'])
        
    if request.method == 'POST':
        if 'nome_empresa' in request.form:
            current_user.nome_empresa = request.form.get('nome_empresa')
            new_email = request.form.get('email')
            
            if new_email != current_user.email and Usuario.query.filter_by(email=new_email).first():
                flash("Erro: Este novo email já está sendo usado por outra conta.", 'danger')
                db.session.rollback()
                return redirect(request.url) 
            
            current_user.email = new_email

            new_password = request.form.get('new_password')
            if new_password:
                current_user.set_password(new_password)

            db.session.commit()
            flash('Informações da conta atualizadas com sucesso!', 'message')
            return redirect(url_for('perfil'))

        if 'logo_file' in request.files:
            file = request.files['logo_file']
            
            if file.filename == '':
                flash('Nenhum arquivo selecionado para upload.', 'danger')
                return redirect(request.url)
                
            if file and allowed_file(file.filename):
                if current_user.logo_path and os.path.exists(current_user.logo_path):
                    os.remove(current_user.logo_path)
                    
                filename = secure_filename(f"{current_user.id}_{file.filename}")
                filepath = os.path.join(app.config['LOGO_FOLDER'], filename)
                file.save(filepath)
                
                current_user.logo_path = filepath
                db.session.commit()
                
                flash('Logo da empresa atualizado com sucesso!', 'message')
                return redirect(url_for('perfil'))
            else:
                flash('Tipo de arquivo não permitido. Use PNG, JPG ou JPEG.', 'danger')
                
    return render_template('perfil.html', user=current_user)

@app.route('/admin/logs')
@login_required
def admin_logs():
    if current_user.email != ADMIN_EMAIL:
        flash("Acesso negado: Somente administradores podem visualizar o log.", 'danger')
        return redirect(url_for('lista_tarefas'))

    logs = LogExecucao.query.order_by(LogExecucao.timestamp.desc()).limit(100).all()
    
    logs_data = []
    for log in logs:
        tarefa = Tarefa.query.get(log.tarefa_id)
        usuario = Usuario.query.get(log.user_id)
        logs_data.append({
            'timestamp': log.timestamp.strftime('%d/%m/%Y %H:%M:%S'),
            'status': log.status,
            'mensagem': log.mensagem,
            'tarefa_nome': tarefa.nome_cliente if tarefa else 'TAREFA DELETADA',
            'empresa_nome': usuario.nome_empresa if usuario else 'USUÁRIO DELETADO'
        })
        
    return render_template('admin_logs.html', logs=logs_data)

@app.route('/planos')
@login_required
def planos():
    plan_data = {
        'TESTE GRATUITO': {'preco': 'R$ 0', 'limite_tarefas': 3, 'limite_execucoes': 5, 'pitch': 'Até 5 execuções de teste para validar o SaaS.', 'color': 'warning', 'is_upgrade': False},
        'BÁSICO': {'preco': 'R$ 149/mês', 'limite_tarefas': 15, 'limite_execucoes': 'Ilimitado', 'pitch': 'Essencial para 15 processos críticos e uso ilimitado.', 'color': 'info', 'is_upgrade': True},
        'MEDIANO': {'preco': 'R$ 399/mês', 'limite_tarefas': 35, 'limite_execucoes': 'Ilimitado', 'pitch': 'Para automação de um departamento inteiro.', 'color': 'primary', 'is_upgrade': True},
        'AVANÇADO': {'preco': 'R$ 699/mês', 'limite_tarefas': 9999, 'limite_execucoes': 'Ilimitado', 'pitch': 'Monitoramento de alto volume (Ilimitado).', 'color': 'success', 'is_upgrade': True},
    }
    
    current_limit = current_user.limite_tarefas
    
    current_plan_name = 'Customizado' 
    for name, data in plan_data.items():
        if data['limite_tarefas'] == current_limit:
            current_plan_name = name
            break
    
    tarefas_count = Tarefa.query.filter_by(user_id=current_user.id).count()
            
    return render_template('planos.html', 
        plans=plan_data, 
        current_limit=current_limit, 
        current_plan_name=current_plan_name, 
        tarefas_count=tarefas_count,
        executions_count=current_user.executions_count
    )


# --- ROTAS PROTEGIDAS (MANTIDAS) ---

@app.route('/')
@login_required 
def lista_tarefas():
    if current_user.limite_tarefas == 3 and current_user.executions_count >= 5:
        flash("Seu limite de 5 execuções gratuitas foi atingido. Adquira um plano para continuar.", 'danger')
        
    tarefas = Tarefa.query.filter_by(user_id=current_user.id).all() 
    jobs = scheduler.get_jobs()
    job_info = {j.id: j.next_run_time.strftime('%d/%m %H:%M') if j.next_run_time else 'N/A' for j in jobs}
    
    log_history = {}
    for tarefa in tarefas:
        logs = LogExecucao.query.filter_by(tarefa_id=tarefa.id) \
                            .order_by(LogExecucao.timestamp.desc()) \
                            .limit(3) \
                            .all()
        log_history[tarefa.id] = logs
        
    user_logo_path = current_user.logo_path if current_user.logo_path and os.path.exists(current_user.logo_path) else None 
    
    return render_template('lista.html', tarefas=tarefas, job_info=job_info, log_history=log_history, user_logo_path=user_logo_path)

@app.route('/nova', methods=['GET', 'POST'])
@login_required 
def nova_tarefa():
    if request.method == 'POST':
        if Tarefa.query.filter_by(user_id=current_user.id).count() >= current_user.limite_tarefas:
            flash(f"Limite de {current_user.limite_tarefas} tarefas atingido para o seu plano.", 'danger')
            return redirect(url_for('lista_tarefas'))

        if current_user.limite_tarefas == 3 and current_user.executions_count >= 5:
            flash("Seu limite de 5 execuções gratuitas foi atingido. Adquira um plano para continuar.", 'danger')
            return redirect(url_for('planos'))

        arquivo_excel = request.files.get('arquivo_excel')
        if not arquivo_excel or not arquivo_excel.filename or not arquivo_excel.filename.endswith('.xlsx'):
            flash('🚨 Erro: Você deve ANEXAR o Arquivo Excel. Por favor, reanexe o arquivo.', 'danger')
            return redirect(url_for('nova_tarefa'))
            
        nome_cliente = request.form.get('nome_cliente')
        hora_agendamento = request.form.get('hora_agendamento')
        email_destino = request.form.get('email_destino')
        telegram_chat_id = request.form.get('telegram_chat_id')
        
        filtro_tipo = request.form.get('filtro_tipo')
        coluna_data = request.form.get('coluna_data')
        dias_alerta = request.form.get('dias_alerta')
        coluna_status = request.form.get('coluna_status')
        palavra_chave = request.form.get('palavra_chave')
        
        coluna_base = request.form.get('coluna_base')
        coluna_acionado = request.form.get('coluna_acionado')
        coluna_atendido = request.form.get('coluna_atendido')
        coluna_cpc = request.form.get('coluna_cpc')
        coluna_acordos = request.form.get('coluna_acordos')
        coluna_pagamentos = request.form.get('coluna_pagamentos')

        if filtro_tipo == 'DATA':
            if not all([coluna_data, dias_alerta]):
                 flash("Erro: Para filtro por DATA, Coluna da Data e Prazo de Alerta são obrigatórios.", 'danger')
                 return redirect(url_for('nova_tarefa'))
        elif filtro_tipo == 'STATUS':
            if not all([coluna_status, palavra_chave]):
                 flash("Erro: Para filtro por STATUS, Coluna e Palavra-Chave são obrigatórios.", 'danger')
                 return redirect(url_for('nova_tarefa'))
        elif filtro_tipo == 'FUNIL':
            if not all([coluna_base, coluna_acionado, coluna_atendido, coluna_cpc, coluna_acordos, coluna_pagamentos]):
                flash("Erro: Para FUNIL, todas as colunas de Batimento (Base, Acionado, Atendido, CPC, Acordos, Pagamentos) são obrigatórias.", 'danger')
                return redirect(url_for('nova_tarefa'))
                
        try:
            if filtro_tipo == 'DATA':
                coluna_data_int = int(coluna_data)
                dias_alerta_int = int(dias_alerta)
                coluna_status_int = None
                palavra_chave = None
            elif filtro_tipo == 'STATUS':
                coluna_status_int = int(coluna_status)
                coluna_data_int = None
                dias_alerta_int = None
            elif filtro_tipo == 'FUNIL':
                coluna_base_int = int(coluna_base)
                coluna_acionado_int = int(coluna_acionado)
                coluna_atendido_int = int(coluna_atendido)
                coluna_cpc_int = int(coluna_cpc)
                coluna_acordos_int = int(coluna_acordos)
                coluna_pagamentos_int = int(coluna_pagamentos)
        except ValueError:
            flash("Erro: Colunas e Dias de Alerta devem ser números inteiros.", 'danger')
            return redirect(url_for('nova_tarefa'))

        nome_seguro = secure_filename(f"{current_user.id}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{arquivo_excel.filename}")
        caminho_arquivo = os.path.join(app.config['UPLOAD_FOLDER'], nome_seguro)
        arquivo_excel.save(caminho_arquivo)
        
        nova_tarefa = Tarefa(
            nome_cliente=nome_cliente,
            caminho_arquivo=caminho_arquivo,
            hora_agendamento=hora_agendamento,
            email_destino=email_destino,
            telegram_chat_id=telegram_chat_id,
            filtro_tipo=filtro_tipo,
            coluna_data=coluna_data_int if filtro_tipo == 'DATA' else None,
            dias_alerta=dias_alerta_int if filtro_tipo == 'DATA' else None,
            coluna_status=coluna_status_int if filtro_tipo == 'STATUS' else None,
            palavra_chave=palavra_chave if filtro_tipo == 'STATUS' else None,
            coluna_base=coluna_base_int if filtro_tipo == 'FUNIL' else None,
            coluna_acionado=coluna_acionado_int if filtro_tipo == 'FUNIL' else None
            ,coluna_atendido=coluna_atendido_int if filtro_tipo == 'FUNIL' else None
            ,coluna_cpc=coluna_cpc_int if filtro_tipo == 'FUNIL' else None
            ,coluna_acordos=coluna_acordos_int if filtro_tipo == 'FUNIL' else None
            ,coluna_pagamentos=coluna_pagamentos_int if filtro_tipo == 'FUNIL' else None
            ,user_id=current_user.id
        )
        
        db.session.add(nova_tarefa)
        db.session.commit()
        
        agendar_tarefa_core(nova_tarefa.id, hora_agendamento)
        
        flash("Nova tarefa de relatório configurada e agendada com sucesso.", 'message')
        return redirect(url_for('lista_tarefas'))

    
    return render_template('nova_tarefa.html')

@app.route('/processar/<int:tarefa_id>')
@login_required
def processar_agora(tarefa_id):
    tarefa = Tarefa.query.filter_by(id=tarefa_id, user_id=current_user.id).first()
    
    if current_user.limite_tarefas == 3 and current_user.executions_count >= 5:
        flash("Seu limite de 5 execuções gratuitas foi atingido. Adquira um plano para continuar.", 'danger')
        return redirect(url_for('planos'))
        
    if tarefa:
        processar_e_enviar_relatorio(tarefa.id) 
        flash("Tarefa executada com sucesso! Verifique seu email.", 'message')
    
    return redirect(url_for('lista_tarefas'))

@app.route('/editar/<int:tarefa_id>', methods=['GET', 'POST'])
@login_required
def editar_tarefa(tarefa_id):
    tarefa = Tarefa.query.filter_by(id=tarefa_id, user_id=current_user.id).first_or_404()

    if request.method == 'POST':
        novo_arquivo = request.files.get('novo_arquivo_excel')
        caminho_antigo = tarefa.caminho_arquivo

        if novo_arquivo and novo_arquivo.filename:
            if novo_arquivo.filename.endswith('.xlsx'):
                novo_nome_seguro = secure_filename(f"{tarefa.id}_REPLACE_{novo_arquivo.filename}")
                novo_caminho = os.path.join(app.config['UPLOAD_FOLDER'], novo_nome_seguro)
                novo_arquivo.save(novo_caminho)

                tarefa.caminho_arquivo = novo_caminho
                if os.path.exists(caminho_antigo):
                    try:
                        os.remove(caminho_antigo)
                    except PermissionError:
                        flash("Erro de Permissão ao atualizar arquivo. Certifique-se de que o arquivo anterior não está aberto em outro programa.", 'danger')
                        db.session.rollback()
                        return redirect(url_for('editar_tarefa', tarefa_id=tarefa.id))
                
                flash("Arquivo Excel atualizado com sucesso.", 'message')
            else:
                flash("Erro: O novo arquivo deve ser no formato .xlsx.", 'danger')
                return redirect(url_for('editar_tarefa', tarefa_id=tarefa.id))
        
        tarefa.nome_cliente = request.form.get('nome_cliente') 
        tarefa.hora_agendamento = request.form.get('hora_agendamento')
        tarefa.email_destino = request.form.get('email_destino')
        
        filtro_tipo = request.form.get('filtro_tipo')
        coluna_data = request.form.get('coluna_data')
        dias_alerta = request.form.get('dias_alerta')
        coluna_status = request.form.get('coluna_status')
        palavra_chave = request.form.get('palavra_chave')
        
        coluna_base = request.form.get('coluna_base')
        coluna_acionado = request.form.get('coluna_acionado')
        coluna_atendido = request.form.get('coluna_atendido')
        coluna_cpc = request.form.get('coluna_cpc')
        coluna_acordos = request.form.get('coluna_acordos')
        coluna_pagamentos = request.form.get('coluna_pagamentos')

        if filtro_tipo == 'DATA':
            if not all([coluna_data, dias_alerta]):
                 flash("Erro: Para filtro por DATA, Coluna da Data e Prazo de Alerta são obrigatórios.", 'danger')
                 return redirect(url_for('editar_tarefa', tarefa_id=tarefa.id))
        elif filtro_tipo == 'STATUS':
             if not all([coluna_status, palavra_chave]):
                 flash("Erro: Para filtro por STATUS, Coluna e Palavra-Chave são obrigatórios.", 'danger')
                 return redirect(url_for('editar_tarefa', tarefa_id=tarefa.id))
        elif filtro_tipo == 'FUNIL':
            if not all([coluna_base, coluna_acionado, coluna_atendido, coluna_cpc, coluna_acordos, coluna_pagamentos]):
                flash("Erro: Para FUNIL, todas as colunas de Batimento são obrigatórias.", 'danger')
                return redirect(url_for('editar_tarefa', tarefa_id=tarefa.id))
        
        try:
            tarefa.filtro_tipo = filtro_tipo
            
            tarefa.coluna_data = int(coluna_data) if filtro_tipo == 'DATA' else None
            tarefa.dias_alerta = int(dias_alerta) if filtro_tipo == 'DATA' else None
            tarefa.coluna_status = int(coluna_status) if filtro_tipo == 'STATUS' else None
            tarefa.palavra_chave = palavra_chave if filtro_tipo == 'STATUS' else None
            
            tarefa.coluna_base = int(coluna_base) if filtro_tipo == 'FUNIL' else None
            tarefa.coluna_acionado = int(coluna_acionado) if filtro_tipo == 'FUNIL' else None
            tarefa.coluna_atendido = int(coluna_atendido) if filtro_tipo == 'FUNIL' else None
            tarefa.coluna_cpc = int(coluna_cpc) if filtro_tipo == 'FUNIL' else None
            tarefa.coluna_acordos = int(coluna_acordos) if filtro_tipo == 'FUNIL' else None
            tarefa.coluna_pagamentos = int(coluna_pagamentos) if filtro_tipo == 'FUNIL' else None
            
        except ValueError:
            flash("Erro: Colunas e Dias de Alerta devem ser números inteiros.", 'danger')
            return redirect(url_for('editar_tarefa', tarefa_id=tarefa.id))

        db.session.commit()
        
        agendar_tarefa_core(tarefa.id, tarefa.hora_agendamento)
        flash("Tarefa atualizada com sucesso.")
        return redirect(url_for('lista_tarefas'))

    return render_template('editar_tarefa.html', tarefa=tarefa)

@app.route('/deletar/<int:tarefa_id>', methods=['POST'])
@login_required
def deletar_tarefa(tarefa_id):
    tarefa = Tarefa.query.filter_by(id=tarefa_id, user_id=current_user.id).first_or_404()
    
    try: scheduler.remove_job(str(tarefa.id))
    except: pass 
    
    if os.path.exists(tarefa.caminho_arquivo):
        try: os.remove(tarefa.caminho_arquivo)
        except PermissionError:
            flash("Erro: O arquivo Excel não foi deletado do disco (Permissão). Exclua manualmente.", 'warning')
            
    LogExecucao.query.filter_by(tarefa_id=tarefa.id).delete()
    db.session.delete(tarefa)
    db.session.commit()
    
    flash("Tarefa deletada com sucesso.")
    return redirect(url_for('lista_tarefas'))


# --- INICIALIZAÇÃO DA APLICAÇÃO (MANTIDA) ---
scheduler = BackgroundScheduler()
scheduler.start()
atexit.register(lambda: scheduler.shutdown())

if __name__ == '__main__':
    with app.app_context():
        if not os.path.exists(app.config['LOGO_FOLDER']):
            os.makedirs(app.config['LOGO_FOLDER'])
        if not os.path.exists(app.config['UPLOAD_FOLDER']):
            os.makedirs(app.config['UPLOAD_FOLDER'])
            
        db.create_all()
        for tarefa in Tarefa.query.all():
            agendar_tarefa_core(tarefa.id, tarefa.hora_agendamento)
            
    print("Servidor Flask inicializado. Acesse http://127.0.0.1:5000/")
    app.run(debug=True, use_reloader=False)